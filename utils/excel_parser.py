# utils/excel_parser.py
import io
import re
from datetime import datetime, date, timedelta
from openpyxl import load_workbook

def limpiar_matricula(valor):
    if valor is None:
        return ""
    texto = str(valor).strip().upper()
    match = re.search(r"R\d{4}[A-Z]{3}", texto)
    return match.group(0) if match else ""

def excel_serial_to_date(value):
    try:
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        if isinstance(value, (int, float)):
            return date(1899, 12, 30) + timedelta(days=int(value))
    except Exception:
        return None
    return None

def extraer_horas(texto):
    if not texto:
        return []
    encontrados = re.findall(r"\b(\d{1,2}:\d{2})(?::\d{2})?\b", str(texto))
    horas = []
    for h in encontrados:
        hh, mm = h.split(":")
        limpia = f"{int(hh):02d}:{mm}"
        if limpia not in horas:
            horas.append(limpia)
    return horas

def excel_time_to_hhmm(value):
    if value in (None, ""):
        return ""
    if isinstance(value, datetime):
        return value.strftime("%H:%M")
    if isinstance(value, (int, float)):
        total_minutes = round(float(value) * 24 * 60)
        hh = (total_minutes // 60) % 24
        mm = total_minutes % 60
        return f"{hh:02d}:{mm:02d}"
    return " / ".join(extraer_horas(value))

def obtener_rgb_celda(celda):
    try:
        fill = celda.fill
        if not fill or not fill.fgColor:
            return ""
        color = fill.fgColor
        if color.type == "rgb" and color.rgb:
            rgb = color.rgb.upper()
            if len(rgb) == 8:
                rgb = rgb[2:]
            return rgb
    except Exception:
        return ""
    return ""

def es_azul(rgb):
    if not rgb or len(rgb) != 6:
        return False
    try:
        r = int(rgb[0:2], 16)
        g = int(rgb[2:4], 16)
        b = int(rgb[4:6], 16)
        return b > 120 and b > r + 25 and b > g + 5
    except Exception:
        return False

def es_rojo(rgb):
    if not rgb or len(rgb) != 6:
        return False
    try:
        r = int(rgb[0:2], 16)
        g = int(rgb[2:4], 16)
        b = int(rgb[4:6], 16)
        return r > 140 and r > g + 25 and r > b + 25
    except Exception:
        return False

def detectar_colores(ws, fila_ini, fila_fin):
    colores = set()
    for row in range(fila_ini, fila_fin + 1):
        for col in range(1, 16):
            rgb = obtener_rgb_celda(ws.cell(row=row, column=col))
            if es_azul(rgb):
                colores.add("COLOR_AZUL_SECO")
            if es_rojo(rgb):
                colores.add("COLOR_ROJO_FRIO")
    return sorted(colores)

def detectar_bloques(ws):
    bloques = []
    for row in range(1, ws.max_row + 1):
        semi = limpiar_matricula(ws.cell(row=row, column=1).value)
        if semi:
            header_row = max(row - 1, 1)
            start = max(header_row, 1)
            siguiente = ws.max_row
            for r2 in range(row + 1, ws.max_row + 1):
                if limpiar_matricula(ws.cell(row=r2, column=1).value):
                    siguiente = r2 - 2
                    break
            end = min(siguiente, row + 12)
            bloques.append({"semi": semi, "semi_row": row, "header_row": header_row, "start": start, "end": end})
    return bloques

def rango_lectura_termica(bloque):
    r = bloque["semi_row"]
    return r, min(r + 3, bloque["end"])

def detectar_envases(ws, fila_ini, fila_fin):
    textos = []
    for row in range(fila_ini, fila_fin + 1):
        for col in range(1, 16):
            valor = ws.cell(row=row, column=col).value
            if valor is not None:
                textos.append(str(valor).strip().upper().replace(" ", ""))
    unidos = " ".join(textos)
    return ("ENVASES" in unidos) or ("RETORNO" in unidos)

def detectar_mercancias(ws, fila_ini, fila_fin, categoria_segmento_fn):
    categorias_ordenadas = []
    marcas_detectadas = set()
    evidencia_congelado = False

    palabras_congelado_real = ["CONGELADO", "-25", "HELADO", "ULTRACONGELADO"]

    for col in range(1, 16):
        for row in range(fila_ini, fila_fin + 1):
            valor = ws.cell(row=row, column=col).value
            if isinstance(valor, str):
                texto_valor = valor.upper()
                if any(p in texto_valor for p in palabras_congelado_real):
                    evidencia_congelado = True

                partes = re.split(r"[/,\n;]+", texto_valor)
                for parte in partes:
                    cat = categoria_segmento_fn(parte)
                    if cat == "CONGELADO_25" and not evidencia_congelado:
                        continue
                    if cat:
                        marcas_detectadas.add(cat)
                        if cat not in categorias_ordenadas:
                            categorias_ordenadas.append(cat)

    colores = detectar_colores(ws, fila_ini, fila_fin)
    if "COLOR_AZUL_SECO" in colores:
        marcas_detectadas.add("COLOR_AZUL_SECO")
        if "SECO" not in categorias_ordenadas:
            categorias_ordenadas.append("SECO")

    if "COLOR_ROJO_FRIO" in colores:
        marcas_detectadas.add("COLOR_ROJO_FRIO")
        if "SECO" not in categorias_ordenadas and "CONGELADO_25" not in categorias_ordenadas:
            if "REFRIGERADO_3" not in categorias_ordenadas:
                categorias_ordenadas.append("REFRIGERADO_3")
                marcas_detectadas.add("REFRIGERADO_3")

    if not evidencia_congelado and "CONGELADO_25" in categorias_ordenadas:
        categorias_ordenadas = [c for c in categorias_ordenadas if c != "CONGELADO_25"]
        marcas_detectadas.discard("CONGELADO_25")

    return categorias_ordenadas, sorted(marcas_detectadas)

def extraer_registros_excel(bytes_archivo, nombre_archivo, es_excel_operativo, obtener_naviera_fn, categoria_segmento_fn, descripcion_termica_fn):
    archivo_memoria = io.BytesIO(bytes_archivo)
    wb = load_workbook(archivo_memoria, data_only=True)

    if "4402" not in wb.sheetnames:
        return [], f"No se encontró la hoja 4402 en {nombre_archivo}"

    ws = wb["4402"]
    registros = []

    for b in detectar_bloques(ws):
        r = b["semi_row"]
        h = b["header_row"]

        puerto = (
            ws.cell(row=r + 2, column=10).value
            or ws.cell(row=r + 2, column=14).value
            or ws.cell(row=r + 1, column=14).value
        )

        naviera = obtener_naviera_fn(puerto)
        horas_fechas = []

        for col in [12, 13]:
            hora_texto = excel_time_to_hhmm(ws.cell(row=h, column=col).value)
            fecha = excel_serial_to_date(ws.cell(row=r, column=col).value)

            for hora in extraer_horas(hora_texto):
                horas_fechas.append({"hora": hora, "fecha": fecha, "col": col})

        if not horas_fechas:
            continue

        termica_ini, termica_fin = rango_lectura_termica(b)
        categorias_ordenadas, marcas_detectadas = detectar_mercancias(ws, termica_ini, termica_fin, categoria_segmento_fn)

        registros.append({
            "Archivo": nombre_archivo,
            "Es Excel operativo": es_excel_operativo,
            "Semi": b["semi"],
            "Puerto": str(puerto).strip() if puerto is not None else "",
            "Naviera": naviera,
            "Horas fechas": horas_fechas,
            "Mercancías detectadas": ", ".join(marcas_detectadas),
            "Orden térmico detectado": " / ".join([c for c in categorias_ordenadas if c != "PDF_POSICIONAL"]),
            "Rango térmico leído": f"{termica_ini}-{termica_fin}",
            "Descripción térmica": descripcion_termica_fn(categorias_ordenadas, marcas_detectadas),
            "Retira envases detectado": "SI" if detectar_envases(ws, b["start"], b["end"]) else "NO",
            "Fila": r,
        })

    return registros, None