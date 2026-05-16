import streamlit as st
from openpyxl import load_workbook
from datetime import datetime, date, timedelta
import pandas as pd
import io
import re

st.set_page_config(page_title="Operativa Melilla", layout="wide")

st.title("🚢 Generador de Operativa Mercadona Melilla")
st.write("Generador de operativa con Excel día anterior, Excel operativo y modo lunes adelantado.")


# =========================================================
# UTILIDADES BÁSICAS
# =========================================================

def excel_serial_to_date(value):
    try:
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        if isinstance(value, (int, float)):
            return date(1899, 12, 30) + timedelta(days=int(value))
    except Exception:
        return None
    return None


def normalizar(texto):
    return str(texto).strip().upper().replace(" ", "")


def limpiar_matricula(valor):
    if valor is None:
        return ""
    texto = str(valor).strip().upper()
    match = re.search(r"R\d{4}[A-Z]{3}", texto)
    return match.group(0) if match else ""


def extraer_horas(texto):
    if not texto:
        return []

    encontrados = re.findall(r"\b(\d{1,2}:\d{2})(?::\d{2})?\b", str(texto))
    horas = []

    for h in encontrados:
        hh, mm = h.split(":")
        limpia = f"{int(hh):02d}:{mm}"
        if limpia not in horas:
            horas.append(limpia)

    return horas


def excel_time_to_hhmm(value):
    if value in (None, ""):
        return ""

    if isinstance(value, datetime):
        return value.strftime("%H:%M")

    if isinstance(value, (int, float)):
        total_minutes = round(float(value) * 24 * 60)
        hh = (total_minutes // 60) % 24
        mm = total_minutes % 60
        return f"{hh:02d}:{mm:02d}"

    horas = extraer_horas(value)
    return " / ".join(horas)


def hora_a_minutos(hora):
    try:
        h, m = hora.split(":")
        return int(h) * 60 + int(m)
    except Exception:
        return 9999


def dia_semana_es(fecha):
    dias = {
        0: "LUNES",
        1: "MARTES",
        2: "MIÉRCOLES",
        3: "JUEVES",
        4: "VIERNES",
        5: "SÁBADO",
        6: "DOMINGO",
    }
    return dias[fecha.weekday()]


def obtener_naviera(puerto):
    p = normalizar(puerto)

    if p in ["PMOTRIL", "P.MOTRIL", "PALMERIA", "P.ALMERIA", "PALMERÍA", "P.ALMERÍA"]:
        return "TRASMEDITERRANEA"

    if p in ["PMALAGA", "P.MALAGA", "PMÁLAGA", "P.MÁLAGA"]:
        return "BALEARIA"

    return "REVISAR"


def llegada_estimada(naviera):
    if naviera == "TRASMEDITERRANEA":
        return "06:30"
    if naviera == "BALEARIA":
        return "21:00"
    return ""


# =========================================================
# COLORES
# =========================================================

def obtener_rgb_celda(celda):
    try:
        fill = celda.fill
        if not fill or not fill.fgColor:
            return ""

        color = fill.fgColor

        if color.type == "rgb" and color.rgb:
            rgb = color.rgb.upper()
            if len(rgb) == 8:
                rgb = rgb[2:]
            return rgb
    except Exception:
        return ""

    return ""


def es_azul(rgb):
    if not rgb or len(rgb) != 6:
        return False

    try:
        r = int(rgb[0:2], 16)
        g = int(rgb[2:4], 16)
        b = int(rgb[4:6], 16)
        return b > 120 and b > r + 25 and b > g + 5
    except Exception:
        return False


def es_rojo(rgb):
    if not rgb or len(rgb) != 6:
        return False

    try:
        r = int(rgb[0:2], 16)
        g = int(rgb[2:4], 16)
        b = int(rgb[4:6], 16)
        return r > 140 and r > g + 25 and r > b + 25
    except Exception:
        return False


def detectar_colores(ws, fila_ini, fila_fin):
    colores = set()

    for row in range(fila_ini, fila_fin + 1):
        for col in range(1, 16):
            celda = ws.cell(row=row, column=col)
            rgb = obtener_rgb_celda(celda)

            if es_azul(rgb):
                colores.add("COLOR_AZUL_SECO")

            if es_rojo(rgb):
                colores.add("COLOR_ROJO_FRIO")

    return sorted(colores)


# =========================================================
# TEMPERATURA Y CARGA
# =========================================================

def categoria_segmento(segmento):
    texto = str(segmento).upper().strip()

    if not texto:
        return None

    if "CONGELADO" in texto:
        return "CONGELADO_25"

    if any(x in texto for x in ["CARNE", "FRUTA", "PESCADO", "REFRIGERADO", "REPESCA"]):
        return "REFRIGERADO_3"

    if (
        "AGUA" in texto
        or "PICKING" in texto
        or "DROGERIA" in texto
        or "DROGUERIA" in texto
        or "COSMETICA" in texto
        or "COSMÉTICA" in texto
        or "ALCOHOL" in texto
        or "FISCAL" in texto
        or re.search(r"\bSP\b", texto)
    ):
        return "SECO"

    return None


def detectar_mercancias(ws, fila_ini, fila_fin):
    categorias_ordenadas = []
    marcas_detectadas = set()

    for col in range(1, 16):
        for row in range(fila_ini, fila_fin + 1):
            valor = ws.cell(row=row, column=col).value

            if isinstance(valor, str):
                texto = valor.upper()

                partes = re.split(r"[/,\n;]+", texto)

                for parte in partes:
                    cat = categoria_segmento(parte)
                    if cat:
                        marcas_detectadas.add(cat)
                        if cat not in categorias_ordenadas:
                            categorias_ordenadas.append(cat)

    colores = detectar_colores(ws, fila_ini, fila_fin)

    if "COLOR_AZUL_SECO" in colores:
        marcas_detectadas.add("COLOR_AZUL_SECO")
        if "SECO" not in categorias_ordenadas:
            categorias_ordenadas.append("SECO")

    if "COLOR_ROJO_FRIO" in colores:
        marcas_detectadas.add("COLOR_ROJO_FRIO")

    return categorias_ordenadas, sorted(marcas_detectadas)


def etiqueta_categoria(cat):
    if cat == "REFRIGERADO_3":
        return "REFRIGERADO 3º"
    if cat == "CONGELADO_25":
        return "-25º"
    if cat == "SECO":
        return "SECO"
    return cat


def descripcion_termica(categorias_ordenadas, marcas_detectadas):
    cats = categorias_ordenadas[:]

    if not cats:
        if "COLOR_ROJO_FRIO" in marcas_detectadas:
            return "REVISAR FRÍO"
        if "COLOR_AZUL_SECO" in marcas_detectadas:
            return "TODO SECO"
        return "REVISAR"

    unicas = list(dict.fromkeys(cats))

    if len(unicas) == 1:
        if unicas[0] == "SECO":
            return "TODO SECO"
        if unicas[0] == "REFRIGERADO_3":
            return "COMPLETO REFRIGERADO 3º"
        if unicas[0] == "CONGELADO_25":
            return "COMPLETO CONGELADO -25º"

    if len(unicas) == 2:
        return f"DELANTERO {etiqueta_categoria(unicas[0])} / TRASERO {etiqueta_categoria(unicas[1])}"

    if len(unicas) >= 3:
        partes = []
        for i, cat in enumerate(unicas):
            etiqueta = etiqueta_categoria(cat)

            if i == 0:
                partes.append(f"DELANTERO {etiqueta}")
            elif i == len(unicas) - 1:
                partes.append(f"TRASERO {etiqueta}")
            else:
                partes.append(etiqueta)

        return " / ".join(partes)

    return "REVISAR"


def es_carga_fria(descripcion):
    desc = str(descripcion).upper()
    return (
        "-25" in desc
        or "REFRIGERADO" in desc
        or "CONGELADO" in desc
        or "3º" in desc
    ) and "TODO SECO" not in desc


def texto_aviso_temperatura(descripcion, completa):
    if completa == "NO" and es_carga_fria(descripcion):
        return f"⚠️ *TEMPERATURA: {descripcion}. Mantener equipo de frío en marcha.*"
    return ""


def detectar_envases(ws, fila_ini, fila_fin):
    textos = []

    for row in range(fila_ini, fila_fin + 1):
        for col in range(1, 16):
            valor = ws.cell(row=row, column=col).value
            if valor is not None:
                textos.append(normalizar(valor))

    unidos = " ".join(textos)
    return ("ENVASES" in unidos) or ("RETORNO" in unidos)


# =========================================================
# LECTURA DE EXCEL
# =========================================================

def detectar_bloques(ws):
    bloques = []

    for row in range(1, ws.max_row + 1):
        semi = limpiar_matricula(ws.cell(row=row, column=1).value)

        if semi:
            header_row = max(row - 1, 1)
            start = max(header_row, 1)

            siguiente = ws.max_row
            for r2 in range(row + 1, ws.max_row + 1):
                if limpiar_matricula(ws.cell(row=r2, column=1).value):
                    siguiente = r2 - 2
                    break

            end = min(siguiente, row + 12)

            bloques.append({
                "semi": semi,
                "semi_row": row,
                "header_row": header_row,
                "start": start,
                "end": end,
            })

    return bloques


def extraer_registros_excel(bytes_archivo, nombre_archivo, es_excel_operativo):
    archivo_memoria = io.BytesIO(bytes_archivo)
    wb = load_workbook(archivo_memoria, data_only=True)

    if "4402" not in wb.sheetnames:
        return [], f"No se encontró la hoja 4402 en {nombre_archivo}"

    ws = wb["4402"]
    bloques = detectar_bloques(ws)

    registros = []

    for b in bloques:
        r = b["semi_row"]
        h = b["header_row"]

        puerto = (
            ws.cell(row=r + 2, column=10).value
            or ws.cell(row=r + 2, column=14).value
            or ws.cell(row=r + 1, column=14).value
        )

        naviera = obtener_naviera(puerto)

        horas_fechas = []

        for col in [12, 13]:
            hora_texto = excel_time_to_hhmm(ws.cell(row=h, column=col).value)
            fecha = excel_serial_to_date(ws.cell(row=r, column=col).value)

            for hora in extraer_horas(hora_texto):
                horas_fechas.append({
                    "hora": hora,
                    "fecha": fecha,
                    "col": col,
                })

        if not horas_fechas:
            continue

        categorias_ordenadas, marcas_detectadas = detectar_mercancias(ws, b["start"], b["end"])
        termica = descripcion_termica(categorias_ordenadas, marcas_detectadas)
        envases = detectar_envases(ws, b["start"], b["end"])

        registros.append({
            "Archivo": nombre_archivo,
            "Es Excel operativo": es_excel_operativo,
            "Semi": b["semi"],
            "Puerto": str(puerto).strip() if puerto is not None else "",
            "Naviera": naviera,
            "Horas fechas": horas_fechas,
            "Mercancías detectadas": ", ".join(marcas_detectadas),
            "Orden térmico detectado": " / ".join(categorias_ordenadas),
            "Descripción térmica": termica,
            "Retira envases detectado": "SI" if envases else "NO",
            "Fila": r,
        })

    return registros, None


# =========================================================
# CONSTRUCCIÓN DE SERVICIOS Y LLEGADAS
# =========================================================

def construir_servicios(registros, fecha_objetivo):
    servicios = []

    for reg in registros:
        horas_objetivo = [
            x for x in reg["Horas fechas"]
            if x["fecha"] == fecha_objetivo
        ]

        if not horas_objetivo:
            continue

        horas_objetivo = sorted(horas_objetivo, key=lambda x: hora_a_minutos(x["hora"]))

        for idx, hf in enumerate(horas_objetivo):
            es_ultima = idx == len(horas_objetivo) - 1

            if len(horas_objetivo) == 1:
                completa = "SI"
                envases = reg["Retira envases detectado"]
            else:
                completa = "SI" if es_ultima else "NO"
                envases = reg["Retira envases detectado"] if es_ultima else "NO"

            servicios.append({
                "Archivo": reg["Archivo"],
                "Semi": reg["Semi"],
                "Puerto": reg["Puerto"],
                "Naviera": reg["Naviera"],
                "Hora": hf["hora"],
                "Fecha": hf["fecha"].strftime("%d/%m/%Y") if hf["fecha"] else "",
                "Descarga completa": completa,
                "Retira envases": envases,
                "Descripción térmica": reg["Descripción térmica"],
                "Mercancías detectadas": reg["Mercancías detectadas"],
                "Orden térmico detectado": reg["Orden térmico detectado"],
            })

    df = pd.DataFrame(servicios)

    if not df.empty:
        df["OrdenHora"] = df["Hora"].apply(hora_a_minutos)
        df = df.sort_values(["OrdenHora", "Semi"]).drop(columns=["OrdenHora"])

    return df


def construir_llegadas(registros, fecha_objetivo, modo):
    llegadas = []

    for reg in registros:
        if not reg["Es Excel operativo"]:
            continue

        fechas = [x["fecha"] for x in reg["Horas fechas"]]
        if fecha_objetivo not in fechas:
            continue

        if reg["Naviera"] == "REVISAR":
            continue

        if modo == "Operativa adelantada lunes (1 Excel)" and reg["Naviera"] == "BALEARIA":
            continue

        llegadas.append({
            "Semi": reg["Semi"],
            "Naviera": reg["Naviera"],
            "Llegada estimada": llegada_estimada(reg["Naviera"]),
            "Descripción térmica": reg["Descripción térmica"],
            "Puerto": reg["Puerto"],
        })

    df = pd.DataFrame(llegadas)

    if not df.empty:
        df = df.drop_duplicates(subset=["Semi", "Naviera"])

    return df


def filtrar_servicios_por_modo(df_servicios, modo):
    if df_servicios.empty:
        return df_servicios

    df = df_servicios.copy()
    df["OrdenHora"] = df["Hora"].apply(hora_a_minutos)

    if modo == "Operativa adelantada lunes (1 Excel)":
        df = df[df["OrdenHora"] <= hora_a_minutos("20:30")]

    if modo == "Completar operativa lunes (2 Excel)":
        df = df[df["OrdenHora"] > hora_a_minutos("20:30")]

    df = df.sort_values(["OrdenHora", "Semi"]).drop(columns=["OrdenHora"])
    return df


# =========================================================
# TEXTO
# =========================================================

def origen_por_defecto(hora):
    if hora == "05:30":
        return "Desde el puerto"
    if hora == "07:00":
        return "A la llegada del buque “06:30h”"
    return "Desde el puerto"


def texto_post_descarga(completa, envases):
    if completa == "NO" and envases == "NO":
        return "*TRAS DESCARGA PARCIAL: DEJEN EL CEPO PUESTO, SAQUEN 📸 y APARQUEN EN EXPLANADAS FRENTE ZONA TALLERES*"

    if completa == "SI" and envases == "NO":
        return (
            "*Informad de BARRAS, SEPARADORES y ESLINGAS* 👍\n\n"
            "*TRAS DESCARGAR: DEJEN EL CEPO JUNTO A LAS ESLINGAS, SAQUEN 📸 y APARQUEN FRENTE AL REGISTRO*"
        )

    if completa == "SI" and envases == "SI":
        return (
            "*Enviad 📸 de la carga e informad de BARRAS, SEPARADORES y ESLINGAS*\n\n"
            "*TRAS DESCARGAR: DEJEN EL CEPO DENTRO EN UN LADO, SAQUEN 📸 y APARQUEN FRENTE AL REGISTRO*"
        )

    return ""


def aviso_final_por_modo(modo):
    if modo == "Operativa adelantada lunes (1 Excel)":
        return "📢 *ESTAD ATENTOS:* Más adelante publicaremos la operativa correspondiente a las llegadas por BALEARIA 20:40h y posteriores entregas, cuando recibamos el archivo actualizado."
    return "📢 *ESTAD ATENTOS:* Más adelante publicaremos la *OPERATIVA DE TTES. NIEVES* asignada, *ADICIONAL* a lo actual."


def generar_texto(fecha_objetivo, df_servicios, df_llegadas, config_servicios, modo):
    texto = ""

    fecha_txt = fecha_objetivo.strftime("%d/%m/%y")
    dia_txt = dia_semana_es(fecha_objetivo)

    if df_llegadas is not None and not df_llegadas.empty:
        for naviera, grupo in df_llegadas.groupby("Naviera"):
            llegada = grupo["Llegada estimada"].iloc[0]

            texto += f"_________*OPERATIVA MERCADONA {dia_txt} {fecha_txt} 05:30H EN TIENDA*_________\n\n"
            texto += f"*HORA ESTIMADA DE LLEGADA DEL 🛳️ LAS {llegada}H ({naviera})*\n\n"

            for _, fila in grupo.iterrows():
                texto += f"{fila['Semi']}_____  {fila['Descripción térmica']}\n"

            texto += "\n_AL NO VENIR TRACTORAS, ESPERAR REMOLQUES A PIE DE BARCO (MACISTAS)._\n\n\n"

    if df_servicios is not None and not df_servicios.empty:
        df_servicios = df_servicios.copy()
        df_servicios["OrdenHora"] = df_servicios["Hora"].apply(hora_a_minutos)
        df_servicios = df_servicios.sort_values(["OrdenHora", "Semi"])

        for hora, grupo in df_servicios.groupby("Hora", sort=False):
            texto += f"---------*OPERATIVA MERCADONA {dia_txt} {fecha_txt} {hora}H EN TIENDA*---------\n\n"
            texto += "⛔️ *Entrada a tienda SOLO a la hora marcada por MERCADONA*\n\n"

            contador = 1

            for idx, fila in grupo.iterrows():
                key = f"{fila['Semi']}_{fila['Hora']}_{idx}"
                cfg = config_servicios.get(key, {})

                if not cfg.get("incluir", True):
                    continue

                chofer = cfg.get("chofer", "SIN ASIGNAR")
                completa = cfg.get("completa", fila["Descarga completa"])
                envases = cfg.get("envases", fila["Retira envases"])
                origen = cfg.get("origen", origen_por_defecto(fila["Hora"]))
                hora_servicio = cfg.get("hora", fila["Hora"])
                observacion = cfg.get("observacion", "").strip()
                termica = cfg.get("termica", fila["Descripción térmica"])

                texto += f"{contador}º {hora_servicio}h _*{chofer.upper()}*_ enganchar el semi *{fila['Semi']}*\n\n"
                texto += f"{origen}\n"
                texto += f"(*{completa}* se descarga completo y *{envases}* retira envases)\n\n"

                aviso_temp = texto_aviso_temperatura(termica, completa)
                if aviso_temp:
                    texto += aviso_temp + "\n\n"

                if observacion:
                    texto += observacion + "\n\n"

                post = texto_post_descarga(completa, envases)
                if post:
                    texto += post + "\n\n"

                contador += 1

            texto += "*----------//----------*\n\n\n"

    texto += aviso_final_por_modo(modo) + "\n\n"
    texto += "*BUEN SERVICIO - GRACIAS - TODOS SOMOS COMPAÑEROS Y TODOS NOS AYUDAMOS*"

    return texto


# =========================================================
# INTERFAZ
# =========================================================

lista_choferes = [
    "Abdel Ali",
    "Abdeslam Zakour",
    "Abdu",
    "Amin",
    "Andrés",
    "Brahim",
    "Chafik",
    "Cucu",
    "Hassan",
    "Ibrahim",
    "Karim Halifa",
    "Mohamed Chilah",
    "Rafa",
    "Reverte",
    "Yassin",
    "Yeray",
    "OTRO",
]

st.subheader("1. Modo de trabajo")

modo = st.selectbox(
    "Selecciona el modo",
    [
        "Operativa completa normal (2 Excel)",
        "Operativa adelantada lunes (1 Excel)",
        "Completar operativa lunes (2 Excel)",
        "Festivo / sin entregas",
    ],
)

fecha_objetivo = st.date_input(
    "Fecha de la operativa",
    value=date.today()
)

if modo == "Festivo / sin entregas":
    st.info("Modo festivo seleccionado. No se genera operativa de entregas.")
    st.stop()

st.subheader("2. Archivos Excel")

archivo_unico = None
archivo_anterior = None
archivo_operativo = None

if modo == "Operativa adelantada lunes (1 Excel)":
    archivo_unico = st.file_uploader(
        "Excel único recibido el domingo para operativa adelantada del lunes",
        type=["xlsm", "xlsx"],
        key="archivo_unico"
    )
else:
    col_a, col_b = st.columns(2)

    with col_a:
        archivo_anterior = st.file_uploader(
            "Excel anterior / ya usado",
            type=["xlsm", "xlsx"],
            key="archivo_anterior"
        )

    with col_b:
        archivo_operativo = st.file_uploader(
            "Excel operativo / actualizado",
            type=["xlsm", "xlsx"],
            key="archivo_operativo"
        )

puede_procesar = False

if modo == "Operativa adelantada lunes (1 Excel)" and archivo_unico:
    puede_procesar = True

if modo != "Operativa adelantada lunes (1 Excel)" and archivo_anterior and archivo_operativo:
    puede_procesar = True

if puede_procesar:
    registros = []

    if modo == "Operativa adelantada lunes (1 Excel)":
        regs, err = extraer_registros_excel(
            archivo_unico.read(),
            archivo_unico.name,
            True
        )

        if err:
            st.error(err)
            st.stop()

        registros.extend(regs)

    else:
        regs_ant, err_ant = extraer_registros_excel(
            archivo_anterior.read(),
            archivo_anterior.name,
            False
        )

        regs_op, err_op = extraer_registros_excel(
            archivo_operativo.read(),
            archivo_operativo.name,
            True
        )

        if err_ant:
            st.error(err_ant)

        if err_op:
            st.error(err_op)

        if err_ant or err_op:
            st.stop()

        registros.extend(regs_ant)
        registros.extend(regs_op)

    df_servicios = construir_servicios(registros, fecha_objetivo)
    df_servicios = filtrar_servicios_por_modo(df_servicios, modo)

    df_llegadas = construir_llegadas(registros, fecha_objetivo, modo)

    if modo == "Completar operativa lunes (2 Excel)" and not df_llegadas.empty:
        df_llegadas = df_llegadas[df_llegadas["Naviera"] == "BALEARIA"]

    st.success("Archivos procesados correctamente.")

    st.subheader("3. Llegadas detectadas")
    if df_llegadas.empty:
        st.warning("No se detectaron llegadas para este modo.")
    else:
        st.dataframe(df_llegadas, use_container_width=True, hide_index=True)

    st.subheader("4. Servicios detectados")
    if df_servicios.empty:
        st.warning("No se detectaron servicios para la fecha seleccionada y el modo actual.")
    else:
        st.dataframe(df_servicios, use_container_width=True, hide_index=True)

        st.subheader("5. Asignación de chóferes y ajustes")
        config_servicios = {}

        for idx, fila in df_servicios.iterrows():
            key = f"{fila['Semi']}_{fila['Hora']}_{idx}"

            st.markdown(f"### {fila['Hora']} — {fila['Semi']}")

            c0, c1, c2, c3, c4 = st.columns([1, 2, 1, 1, 3])

            with c0:
                incluir = st.checkbox(
                    "Incluir",
                    value=True,
                    key=f"incluir_{key}"
                )

            with c1:
                chofer_sel = st.selectbox(
                    "Chófer",
                    lista_choferes,
                    key=f"chofer_{key}"
                )

                if chofer_sel == "OTRO":
                    chofer = st.text_input(
                        "Escribir chófer",
                        key=f"chofer_otro_{key}"
                    )
                else:
                    chofer = chofer_sel

            with c2:
                completa = st.selectbox(
                    "Completa",
                    ["SI", "NO"],
                    index=0 if fila["Descarga completa"] == "SI" else 1,
                    key=f"completa_{key}"
                )

            with c3:
                envases = st.selectbox(
                    "Envases",
                    ["SI", "NO"],
                    index=0 if fila["Retira envases"] == "SI" else 1,
                    key=f"envases_{key}"
                )

            with c4:
                origen = st.text_input(
                    "Origen / instrucción",
                    value=origen_por_defecto(fila["Hora"]),
                    key=f"origen_{key}"
                )

            termica_manual = st.text_input(
                "Descripción térmica / temperatura",
                value=fila["Descripción térmica"],
                key=f"termica_{key}"
            )

            observacion = st.text_area(
                "Observación adicional para este servicio",
                value="",
                height=70,
                key=f"obs_{key}"
            )

            config_servicios[key] = {
                "incluir": incluir,
                "chofer": chofer,
                "completa": completa,
                "envases": envases,
                "origen": origen,
                "hora": fila["Hora"],
                "observacion": observacion,
                "termica": termica_manual,
            }

        if st.button("🚀 GENERAR OPERATIVA"):
            texto = generar_texto(
                fecha_objetivo,
                df_servicios,
                df_llegadas,
                config_servicios,
                modo
            )

            st.subheader("Operativa generada")
            st.text_area("Texto listo para copiar", texto, height=700)